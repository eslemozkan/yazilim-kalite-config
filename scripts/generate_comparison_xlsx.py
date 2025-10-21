#!/usr/bin/env python3
"""
Karşılaştırma Excel Raporu Oluşturucu
Yazılım Kalite ve Güvence - Konfigürasyon/Güvenlik Başlıkları Testi

Bu script, ham JSON raporlarını okuyup karşılaştırma tablosu oluşturur.

Kullanım:
    python scripts/generate_comparison_xlsx.py
    python scripts/generate_comparison_xlsx.py --input data/raw_reports/
"""

import json
import argparse
import os
import sys
from datetime import datetime
from typing import Dict, List, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class ComparisonGenerator:
    """Karşılaştırma Excel raporu oluşturucu sınıfı"""
    
    def __init__(self):
        self.data = []
        self.headers = ['Target', 'Header_Name', 'Value', 'Status', 'Severity', 'Remark']
    
    def load_json_files(self, input_path: str) -> List[Dict]:
        """JSON dosyalarını yükler"""
        data = []
        
        if os.path.isfile(input_path):
            # Tek dosya
            with open(input_path, 'r', encoding='utf-8') as f:
                file_data = json.load(f)
                if isinstance(file_data, list):
                    data.extend(file_data)
                else:
                    data.append(file_data)
        else:
            # Dizin
            if not os.path.exists(input_path):
                print(f"Directory not found: {input_path}")
                return data
            
            for filename in os.listdir(input_path):
                if filename.endswith('.json') and not filename.startswith('all_'):
                    file_path = os.path.join(input_path, filename)
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            file_data = json.load(f)
                            if isinstance(file_data, list):
                                data.extend(file_data)
                            else:
                                data.append(file_data)
                    except Exception as e:
                        print(f"Error loading {filename}: {str(e)}")
        
        return data
    
    def create_comparison_data(self, data: List[Dict]) -> List[Dict]:
        """Karşılaştırma verisi oluşturur"""
        comparison_data = []
        
        for target_data in data:
            target_name = target_data.get('target', 'unknown')
            findings = target_data.get('findings', [])
            
            for finding in findings:
                row = {
                    'Target': target_name,
                    'Header_Name': finding.get('name', ''),
                    'Value': finding.get('value', ''),
                    'Status': finding.get('status', ''),
                    'Severity': finding.get('severity', ''),
                    'Remark': finding.get('remark', '')
                }
                comparison_data.append(row)
        
        return comparison_data
    
    def create_excel_report(self, comparison_data: List[Dict], output_file: str) -> None:
        """Excel raporu oluşturur"""
        # Workbook oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Security Headers Comparison"
        
        # Stil tanımları
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Severity renkleri
        severity_colors = {
            'High': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            'Medium': PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
            'Low': PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
        }
        
        # Status renkleri
        status_colors = {
            'pass': PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
            'warn': PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
            'fail': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        }
        
        # Başlık satırı
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Veri satırları
        for row_idx, row_data in enumerate(comparison_data, 2):
            for col_idx, header in enumerate(self.headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Severity rengi
                if header == 'Severity':
                    severity = row_data[header]
                    if severity in severity_colors:
                        cell.fill = severity_colors[severity]
                
                # Status rengi
                if header == 'Status':
                    status = row_data[header]
                    if status in status_colors:
                        cell.fill = status_colors[status]
        
        # Sütun genişliklerini ayarla
        column_widths = {
            'A': 15,  # Target
            'B': 25,  # Header_Name
            'C': 40,  # Value
            'D': 10,  # Status
            'E': 12,  # Severity
            'F': 50   # Remark
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Filtre ekle
        ws.auto_filter.ref = f"A1:{chr(65 + len(self.headers) - 1)}{len(comparison_data) + 1}"
        
        # Özet sayfası oluştur
        self.create_summary_sheet(wb, comparison_data)
        
        # Dosyayı kaydet
        wb.save(output_file)
        print(f"Excel report saved: {output_file}")
    
    def create_summary_sheet(self, wb: Workbook, comparison_data: List[Dict]) -> None:
        """Özet sayfası oluşturur"""
        ws = wb.create_sheet("Summary")
        
        # Veri analizi
        targets = set(row['Target'] for row in comparison_data)
        severities = {}
        statuses = {}
        header_stats = {}
        
        for row in comparison_data:
            target = row['Target']
            severity = row['Severity']
            status = row['Status']
            header = row['Header_Name']
            
            # Severity sayımı
            if severity not in severities:
                severities[severity] = 0
            severities[severity] += 1
            
            # Status sayımı
            if status not in statuses:
                statuses[status] = 0
            statuses[status] += 1
            
            # Header istatistikleri
            if header not in header_stats:
                header_stats[header] = {'total': 0, 'pass': 0, 'warn': 0, 'fail': 0}
            header_stats[header]['total'] += 1
            header_stats[header][status] += 1
        
        # Özet bilgileri
        summary_data = [
            ["SUMMARY REPORT", ""],
            ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Total Targets", len(targets)],
            ["Total Findings", len(comparison_data)],
            ["", ""],
            ["SEVERITY DISTRIBUTION", ""],
        ]
        
        for severity, count in sorted(severities.items()):
            percentage = (count / len(comparison_data) * 100) if comparison_data else 0
            summary_data.append([severity, f"{count} ({percentage:.1f}%)"])
        
        summary_data.extend([
            ["", ""],
            ["STATUS DISTRIBUTION", ""],
        ])
        
        for status, count in sorted(statuses.items()):
            percentage = (count / len(comparison_data) * 100) if comparison_data else 0
            summary_data.append([status.upper(), f"{count} ({percentage:.1f}%)"])
        
        summary_data.extend([
            ["", ""],
            ["HEADER STATISTICS", ""],
            ["Header Name", "Total", "Pass", "Warn", "Fail"],
        ])
        
        for header, stats in sorted(header_stats.items()):
            summary_data.append([
                header,
                stats['total'],
                stats['pass'],
                stats['warn'],
                stats['fail']
            ])
        
        # Veriyi sayfaya yaz
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Başlık
                    cell.font = Font(bold=True, size=14)
                elif row_idx in [6, 12, 18]:  # Alt başlıklar
                    cell.font = Font(bold=True)
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
    
    def run_generation(self, input_path: str = 'data/raw_reports', output_dir: str = 'data/processed') -> None:
        """Karşılaştırma raporu oluşturur"""
        # Output dizinini oluştur
        os.makedirs(output_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # JSON dosyalarını yükle
        data = self.load_json_files(input_path)
        
        if not data:
            print("No data found to generate comparison")
            return
        
        # Karşılaştırma verisi oluştur
        comparison_data = self.create_comparison_data(data)
        
        # Excel raporu oluştur
        output_file = f"{output_dir}/comparison_table_{timestamp}.xlsx"
        self.create_excel_report(comparison_data, output_file)
        
        print(f"Comparison report generated with {len(comparison_data)} findings")

def main():
    """Ana fonksiyon"""
    parser = argparse.ArgumentParser(description='Generate comparison Excel report')
    parser.add_argument('--input', default='data/raw_reports',
                       help='Input directory or JSON file')
    parser.add_argument('--output', default='data/processed',
                       help='Output directory for Excel report')
    
    args = parser.parse_args()
    
    # Generator'ı başlat
    generator = ComparisonGenerator()
    
    # Raporu oluştur
    generator.run_generation(args.input, args.output)

if __name__ == '__main__':
    main()
